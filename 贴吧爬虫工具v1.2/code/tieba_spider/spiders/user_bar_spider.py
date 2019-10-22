import re
import os
import scrapy
import json
import urllib
from openpyxl import load_workbook
from tieba_spider.items import UserBarItem
import tieba_spider
import tkinter as tk



# 爬取用户的关注信息
class UserBarSpider(scrapy.Spider):
    name = 'user_bar_spider'
    print('user_bar_spider的当前路径是：' + os.getcwd())
    start_urls = [
        'https://tieba.baidu.com/home/main?un=%E4%BA%BA%E6%A0%BC%E5%88%86%E8%A3%82%E7%9A%84%E5%A6%96%E5%AD%BD&lp=home_main_follow_main']

    def __init__(self, config=None, *args, **kwargsf):
        if not config:
            pass
        try:
            if config:
                tieba_spider.config = config[0]  # 将字符串 转化为字典
                tieba_spider.config['bar_name'] = urllib.parse.unquote(tieba_spider.config['bar_name'])
                tieba_spider.config['post_num'] = int(tieba_spider.config['post_num'])
                tieba_spider.config['spider_name'] = self.name
        except TypeError:
            pass
        self.check_path()
        if tieba_spider.config['bar_name']:
            self.bar_info_wb = load_workbook('%s%s.xlsx' % (tieba_spider.excel_path, tieba_spider.config['bar_name']))
        else:
            tieba_spider.config['bar_name'] = input('请输入要读取用户的Excel文件:')
            self.bar_info_wb = load_workbook('%s%s.xlsx' % (tieba_spider.excel_path, tieba_spider.config['bar_name']))
        self.user_sheet = self.bar_info_wb['用户']


    def check_path(self):
        print('当前路径是：' + os.getcwd())
        if re.search('dist', os.getcwd()):
            tieba_spider.excel_path = '../../../%s/' % tieba_spider.config['bar_name']
            # 查看有无Excel文件夹  exe
            if os.path.exists(tieba_spider.excel_path):
                print('文件夹已存在')
            # 创建文件夹
            else:
                os.mkdir(tieba_spider.excel_path)

    def parse(self, response):
        # 爬取用户关注贴吧的信息  移动端
        headers = {
            'user-agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 6_0 like Mac OS X) AppleWebKit/536.26 (KHTML, like Gecko) Version/6.0 Mobile/10A5376e Safari/8536.25'}

        for url_cell in list(self.user_sheet.columns)[1][1:]:  # 第二列 第二行开始
            yield scrapy.Request(
                url_cell.value,
                headers=headers,
                callback=self.parse_user_bar,
                dont_filter=True,
            )
        self.bar_info_wb.close()

    def parse_user_bar(self, response):
        user_bar_item = UserBarItem()
        has_hide = 0  # 默认用户没有屏蔽动态
        user_bar_item['url'] = response.request.url  # 用户链接
        user_bar_item['name'] = response.xpath("//*[@id='i_head']/div/a/span[2]/text()").extract_first()
        user_bar_item['bar_num'] = response.xpath("/html/body/div[4]/a[2]/span[1]/text()").extract_first()
        if response.xpath("//div[@class='home_concern_forum_title']"):
            for li in response.xpath("//div[@class='home_concern_forum_title']"):
                user_bar_item['bar_name'] = li.xpath("./text()").extract_first()
                user_bar_item['bar_name'] = user_bar_item['bar_name'].rstrip()  # 去除抓取到的\xa0 空格字符
                user_bar_item['bar_grade'] = li.xpath("./span/text()").extract_first()
                yield user_bar_item
        else:
            has_hide = 1  # 用户屏蔽了动态

        # # 需要解析ajax的json内容
        if int(user_bar_item['bar_num']) > 10 and has_hide == 0:
            ajax_url = ('https://tieba.baidu.com/home/concern?un=%s' % urllib.parse.quote(user_bar_item['name']) +
                        '&is_ajax=1&lp=home_main_concern_more&pn=2')
            # 爬取用户关注贴吧的信息  移动端
            headers = {
                'user-agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 6_0 like Mac OS X) AppleWebKit/536.26 (KHTML, like Gecko) Version/6.0 Mobile/10A5376e Safari/8536.25'}
            yield scrapy.Request(
                ajax_url,
                headers=headers,
                callback=self.parse_bar_ajax,
                meta={"user_bar_item": user_bar_item},
                dont_filter=True,
            )

    # 抓取ajax加载的信息
    def parse_bar_ajax(self, response):
        user_bar_item = response.meta['user_bar_item']
        data = json.loads(response.body)
        code = data['data']['content']
        has_more = data['data']['page']['has_more']  # int
        bar_name_list = re.findall(r"home_concern_forum_title.>(.*?)&", code)  # 找吧名
        bar_grade_list = re.findall(r'level_(.*?)"', code)  # 找等级
        for i in range(len(bar_name_list)):
            user_bar_item['bar_name'] = bar_name_list[i]
            user_bar_item['bar_grade'] = bar_grade_list[i]
            yield user_bar_item

        if has_more:
            headers = {
                'user-agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 6_0 like Mac OS X) AppleWebKit/536.26 (KHTML, like Gecko) Version/6.0 Mobile/10A5376e Safari/8536.25'}
            page = int(response.url.split('pn=')[1]) + 1
            next_url = response.url.split('pn=')[0] + 'pn=' + repr(page)
            yield scrapy.Request(
                next_url,
                headers=headers,
                callback=self.parse_bar_ajax_more,
                meta={"user_bar_item": user_bar_item},
                dont_filter=True,
            )

    # 抓取ajax加载的信息
    def parse_bar_ajax_more(self, response):
        user_bar_item = response.meta['user_bar_item']
        data = json.loads(response.body)
        code = data['data']['content']
        has_more = data['data']['page']['has_more']  # int
        bar_name_list = re.findall(r"home_concern_forum_title.>(.*?)&", code)  # 找吧名
        bar_grade_list = re.findall(r'level_(.*?)"', code)  # 找等级
        for i in range(len(bar_name_list)):
            user_bar_item['bar_name'] = bar_name_list[i]
            user_bar_item['bar_grade'] = bar_grade_list[i]
            yield user_bar_item

        if has_more:
            headers = {
                'user-agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 6_0 like Mac OS X) AppleWebKit/536.26 (KHTML, like Gecko) Version/6.0 Mobile/10A5376e Safari/8536.25'}
            page = int(response.url.split('pn=')[1]) + 1
            next_url = response.url.split('pn=')[0] + 'pn=' + repr(page)
            yield scrapy.Request(
                next_url,
                headers=headers,
                callback=self.parse_bar_ajax,
                meta={"user_bar_item": user_bar_item},
                dont_filter=True,
            )


from scrapy.crawler import CrawlerProcess
from scrapy.utils.project import get_project_settings

if __name__ == '__main__':
    process = CrawlerProcess(settings=get_project_settings())
    process.crawl('user_bar_spider')
    process.start()

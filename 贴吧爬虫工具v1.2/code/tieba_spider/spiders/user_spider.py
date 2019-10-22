import scrapy
import re
import json
import urllib
import time
from openpyxl import load_workbook
from tieba_spider.items import UserItem, NumItem
import tieba_spider
import os
import tieba_spider.pipelines
from scrapy.downloadermiddlewares.retry import RetryMiddleware
import sys


class UserSpiderSpider(scrapy.Spider):
    name = 'user_spider'
    start_urls = ['https://tieba.baidu.com/f?kw=%E4%B8%89%E5%9B%BD%E6%9D%80']  # 三国杀吧 PC贴吧
    print('user_spider的当前路径是：' + os.getcwd())

    def __init__(self, config=None, *args, **kwargsf):
        if not config:
            pass
        try:
            if config:
                tieba_spider.config = config[0]  # 将字符串 转化为字典
                tieba_spider.config['bar_name'] = urllib.parse.unquote(tieba_spider.config['bar_name'])
                tieba_spider.config['post_num'] = int(tieba_spider.config['post_num'])
                tieba_spider.config['spider_name'] = self.name
        except TypeError:
            pass
        self.check_path()
        if tieba_spider.config['bar_name']:
            self.bar_info_wb = load_workbook('%s%s.xlsx' % (tieba_spider.excel_path, tieba_spider.config['bar_name']))
        else:
            # tieba_spider.config['bar_name'] = input('请输入要读取的Excel文件:')
            tieba_spider.config['bar_name'] = '三国杀'
            self.bar_info_wb = load_workbook('%s%s.xlsx' % (tieba_spider.excel_path, tieba_spider.config['bar_name']))
        self.bar_info_sheet = self.bar_info_wb['帖子']

    def check_path(self):
        print('当前路径是：' + os.getcwd())
        if re.search('dist', os.getcwd()):
            tieba_spider.excel_path = '../../../%s/' % tieba_spider.config['bar_name']
            # 查看有无Excel文件夹  exe
            if os.path.exists(tieba_spider.excel_path):
                print('文件夹已存在')
            # 创建文件夹
            else:
                os.mkdir(tieba_spider.excel_path)

    def parse(self, response):
        for url_cell in list(self.bar_info_sheet.columns)[1][1:]:  # 第二列 第二行开始
            num_item = NumItem()
            num_item['id'] = []
            num_item['row'] = url_cell.row
            num_item['has_crawl_page'] = 0
            yield scrapy.Request(
                url_cell.value,
                callback=self.parse_detail,
                meta={"num_item": num_item}  # 参与的用户数量
            )
        self.bar_info_wb.close()

    def parse_detail(self, response):
        num_item = response.meta['num_item']
        page_comment_has_crawl = 0
        is_end_page = 0

        # 判断当前页面是否是最后一页
        if response.xpath("//span[@class='tP']/text()").extract_first():  # 第几页
            page = response.xpath("//span[@class='tP']/text()").extract_first()
        else:
            page = '1'
        end_page = response.xpath("//*[@id='thread_theme_5']/div[1]/ul/li[2]/span[2]/text()").extract_first()
        num_item['total_page'] = end_page  # str
        if page == end_page:
            is_end_page = 1  # 统计人数
        # 获取用户链接
        for floor in response.xpath("//*[@id='j_p_postlist']/div"):
            # 得到用户网页链接
            data = json.loads(floor.xpath("@data-field").extract_first())
            user_item = UserItem()
            if data['author']['user_id'] not in num_item['id']:
                num_item['id'].append(data['author']['user_id'])  # 用户的唯一ID
            if data['author']['user_name']:
                user_item['name'] = data['author']['user_name']
            else:
                user_item['name'] = data['author']['user_nickname']
            user_item['url'] = 'https://tieba.baidu.com/home/main?un=%s' % user_item['name']  # 有可能重名
            # 抓取客户端信息 由于不同的吧 信息标签不同  使用正则表达式兼容性比较强
            try:
                user_item['client'] = data['content']['open_type']
            except KeyError:
                user_item['client'] = floor.xpath(".//span[@class='tail-info']//a/text()").extract_first()
                tieba_spider.config['bar_type'] = 'old_school'
            else:
                if user_item['client'] == 'apple':
                    user_item['client'] = 'iPhone客户端'
                elif user_item['client'] == 'android':
                    user_item['client'] = 'Android客户端'
                tieba_spider.config['bar_type'] = 'new_school'

            if not user_item['client']:
                user_item['client'] = 'PC'

            yield scrapy.Request(  # 进入用户的主界面
                user_item['url'],
                callback=self.parse_user,
                meta={"user_item": user_item},
            )

            if data['content']['comment_num'] > 0 and not page_comment_has_crawl:  # 该楼层有用户在讨论，需要爬取
                if tieba_spider.config['bar_type'] == 'old_school':
                    page_comment_has_crawl = 1  # 一页只需要爬取一次评论
                    t = int(round(time.time()) * 1000)  # 时间戳
                    tid = data['content']['thread_id']  # 帖子ID
                    fid = data['content']['forum_id']  # 贴吧ID
                    # comment_url = 'http://dq.tieba.com/p/totalComment?t=%d&tid=%s&fid=%s&pn=%s&see_lz=0' % (
                    #     t, tid, fid, page)
                    comment_url = 'http://tieba.com/p/totalComment?t=%d&tid=%s&fid=%s&pn=%s&see_lz=0' % (
                        t, tid, fid, page)
                    # 爬取评论用户
                    # 华为吧没有total_comment
                    yield scrapy.Request(
                        comment_url,
                        callback=self.parse_comment,
                        meta={"num_item": num_item, "is_end_page": is_end_page},
                        dont_filter=True,
                    )
                elif tieba_spider.config['bar_type'] == 'new_school':
                    pass
        if page_comment_has_crawl == 0:  # 这一页没有楼中楼
            num_item['has_crawl_page'] += 1
        # 自动翻页
        next_url = response.xpath(
            "//li[@class='l_pager pager_theme_4 pb_list_pager']//a[text()='下一页']/@href").extract_first()
        if next_url is not None:
            next_url = urllib.parse.urljoin(response.url, next_url)
            yield scrapy.Request(
                next_url,
                callback=self.parse_detail,
                meta={"num_item": num_item}
            )
        # 当此页面是最后是最后一页而且没有用户评论的时候执行
        if int(num_item['total_page']) == num_item['has_crawl_page'] and next_url is None:
            yield num_item

    def parse_comment(self, response):
        is_end_page = response.meta['is_end_page']
        num_item = response.meta['num_item']
        num_item['has_crawl_page'] += 1
        try:
            data = json.loads(response.body)
        except:
            print('用户列表json文件加载失败，请检查请求url是否正确')
        user_list = data['data']['user_list']
        for i in user_list:
            if user_list[i]['user_id'] not in num_item['id']:
                num_item['id'].append(user_list[i]['user_id'])  # 用户的唯一ID
                try:
                    user_url = 'https://tieba.baidu.com/home/main?un=%s' % user_list[i]['user_name']  # 有可能重名
                except KeyError:
                    try:
                        user_url = 'https://tieba.baidu.com/home/main?un=%s' % user_list[i]['user_nickname']  # 有可能重名
                    except KeyError:
                        print('用户资料出错')  # 有些用户只有ID 没有名字
                        continue
                # 进入用户的界面
                yield scrapy.Request(  # 进入用户的主界面
                    user_url,
                    callback=self.parse_comment_user,
                    # dont_filter=True,
                )

        if int(num_item['total_page']) == num_item['has_crawl_page']:
            yield num_item

    def parse_user(self, response):
        user_item = response.meta['user_item']
        user_item['url'] = response.request.url  # 用户链接
        if response.xpath("//span[@class='userinfo_sex userinfo_sex_male']"):
            user_item['sex'] = '男'
        else:
            user_item['sex'] = '女'
        user_item['focus_num'] = response.xpath("//span[@class='concern_num']/a[1]/text()").extract_first()
        user_item['fan_num'] = response.xpath("//*[@id='container']/div[2]/div[3]/h1/span/a/text()").extract_first()
        # 不知道上面原因吧龄，和发帖数量不能直接提取文本，所以要从字符串里面提取
        str = response.xpath("//span[@class='user_name']").extract_first()
        user_item['year'] = re.findall(r"吧龄:(.*?)[年<]", str)[0]
        user_item['post_num'] = re.findall(r":(.*?)<", str)[2]
        yield user_item

    def parse_comment_user(self, response):
        user_item = UserItem()
        user_item['url'] = response.request.url  # 用户链接
        try:
            user_item['name'] = response.xpath("//span[@class='user_name']/text()").extract_first()
            user_item['name'] = user_item['name'].split(':')[1]
        except AttributeError:
            user_item['name'] = response.xpath("//span[@class='userinfo_username ']/text()").extract_first()
        user_item['client'] = '楼层讨论用户'
        if response.xpath("//span[@class='userinfo_sex userinfo_sex_male']"):
            user_item['sex'] = '男'
        else:
            user_item['sex'] = '女'
        user_item['focus_num'] = response.xpath("//span[@class='concern_num']/a[1]/text()").extract_first()
        user_item['fan_num'] = response.xpath("//*[@id='container']/div[2]/div[3]/h1/span/a/text()").extract_first()
        # 不知道上面原因吧龄，和发帖数量不能直接提取文本，所以要从字符串里面提取
        str = response.xpath("//span[@class='user_name']").extract_first()
        user_item['year'] = re.findall(r"吧龄:(.*?)[年<]", str)[0]
        user_item['post_num'] = re.findall(r":(.*?)<", str)[2]
        yield user_item


from scrapy.crawler import CrawlerProcess
from scrapy.utils.project import get_project_settings

if __name__ == '__main__':
    process = CrawlerProcess(settings=get_project_settings())
    process.crawl('user_spider')
    process.start()

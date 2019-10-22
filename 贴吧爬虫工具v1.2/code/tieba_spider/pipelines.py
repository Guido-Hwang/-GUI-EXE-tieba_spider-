import os
import os.path
import re
from openpyxl import Workbook  # 导入Excel
from openpyxl import load_workbook
from tieba_spider.items import PostItem, PostContentItem, UserItem, UserBarItem, NumItem
import tieba_spider
import time
import middle_file


class BarSpiderPipeline(object):  # 贴吧信息处理管道
    def __init__(self, spider_data):
        if tieba_spider.config['comment_crawl_flag']:
            file_name = '%s%s.txt' % (tieba_spider.excel_path, tieba_spider.config['bar_name'])
            self.comment_content_file = open(file_name, 'w')
            self.commnent_count = 0
        else:
            self.check_path()
            self._connect_excel()
        self.spider_data = spider_data

    @classmethod
    def from_crawler(cls, crawler):
        """
        获取spider的settings参数,返回Pipeline实例对象
        """
        spider_data = crawler.settings.get("SPIDER_DATA")
        print("### pipeline get spider_data: {}".format(spider_data))

        return cls(spider_data)

    def close_spider(self, spider):
        print("### spdier close {}".format(spider.name))
        if tieba_spider.config['comment_crawl_flag']:
            self.comment_content_file.close()
        else:
            self.bar_wb.save('%s%s.xlsx' % (tieba_spider.excel_path, tieba_spider.config['bar_name']))

    def check_path(self):
        print('当前路径是：' + os.getcwd())
        if re.search('dist', os.getcwd()):
            tieba_spider.excel_path = '../../../%s/' % tieba_spider.config['bar_name']
            # 查看有无Excel文件夹  exe
            if os.path.exists(tieba_spider.excel_path):
                print('文件夹已存在')
            # 创建文件夹
            else:
                os.mkdir(tieba_spider.excel_path)

    def _connect_excel(self):
        file_name = '%s%s.xlsx' % (tieba_spider.excel_path, tieba_spider.config['bar_name'])
        if os.path.isfile(file_name):
            print('Excel已存在')
            self.bar_wb = load_workbook('%s%s.xlsx' % (tieba_spider.excel_path, tieba_spider.config['bar_name']))
            self.post_sheet = self.bar_wb['帖子']
            self.user_sheet = self.bar_wb['用户']
            self.user_bar_sheet = self.bar_wb['用户贴吧']
        else:
            # 创建Excel，编写表头
            print('创建Excel表')
            self.bar_wb = Workbook()
            self.post_sheet = self.bar_wb.create_sheet('帖子', 0)
            self.post_sheet.append(['帖名', 'url', '开贴时间', '回复数量', '用户数量'])
            self.user_sheet = self.bar_wb.create_sheet('用户', 1)
            self.user_sheet.append(['用户名', 'url', '性别', '吧龄', '发帖数量', '关注数量', '粉丝数量', '客户端'])
            self.user_bar_sheet = self.bar_wb.create_sheet('用户贴吧', 2)
            self.user_bar_sheet.append(['用户名', 'url', '贴吧数量', '吧名', '等级'])
            self.bar_wb.save('%s%s.xlsx' % (tieba_spider.excel_path, tieba_spider.config['bar_name']))

    def process_item(self, item, spider):
        if isinstance(item, PostContentItem):  # 评论内容
            if not tieba_spider.comment_output_lock:
                self.comment_content_file.write(item['comment_content'] + '\n')
                print(time.strftime("PostContentitem inputed the data %Y-%m-%d %H:%M:%S", time.localtime()))
                self.commnent_count += 1
                if self.commnent_count > tieba_spider.config['comment_num']:
                    tieba_spider.comment_output_lock = True  # 上锁 防止另外并行的请求多起启动爬虫
                    spider.crawler.engine.close_spider(spider, '超过%d条评论，停止爬取' % tieba_spider.config['comment_num'])
        elif isinstance(item, PostItem):  # 帖子信息
            if not tieba_spider.post_output_lock:
                line = [item['title'], item['url'], item['post_create_time'], item['reply_num']]
                self.post_sheet.append(line)
                print(time.strftime("Postitem inputed the data %Y-%m-%d %H:%M:%S", time.localtime()))
                if self.post_sheet.max_row > tieba_spider.config['post_num']:
                    tieba_spider.post_output_lock = True  # 上锁 防止另外并行的请求多起启动爬虫
                    spider.crawler.engine.close_spider(spider, '超过%d个帖子，停止爬取' % tieba_spider.config['post_num'])
        elif isinstance(item, UserItem):
            if not tieba_spider.user_output_lock:
                # 插入Excel
                line = [item['name'], item['url'], item['sex'], item['year'], item['post_num'], item['focus_num'],
                        item['fan_num'], item['client']]
                self.user_sheet.append(line)
                print(time.strftime("UserItem inputed the data %Y-%m-%d %H:%M:%S", time.localtime()))
                # 没开启统计功能
                if self.user_sheet.max_row > tieba_spider.config['user_num'] and not tieba_spider.config['count_flag']:
                    tieba_spider.user_output_lock = True  # 上锁 防止另外并行的请求多起启动爬虫
                    spider.crawler.engine.close_spider(spider, '超过%d个用户，停止爬取' % tieba_spider.config['user_num'])
        elif isinstance(item, UserBarItem):
            line = [item['name'], item['url'], item['bar_num'], item['bar_name'], item['bar_grade']]
            self.user_bar_sheet.append(line)
            print(time.strftime("UserBarItem inputed the data %Y-%m-%d %H:%M:%S", time.localtime()))
        elif isinstance(item, NumItem):  # 插入用户数量
            if tieba_spider.config['count_flag']:
                print('该贴的讨论人数:', len(item['id']))
                location = 'E' + str(item['row'])
                self.post_sheet[location] = len(item['id'])

        if middle_file.config['spider_status'] == 2:
            middle_file.config['spider_status'] = 1
            spider.crawler.engine.close_spider(spider, '用户主动结束爬虫')
        elif middle_file.config['spider_status'] == 3:
            self.bar_wb.save('%s%s.xlsx' % (tieba_spider.excel_path, tieba_spider.config['bar_name']))
            self.pause_func()

        return item

    def pause_func(self):
        while middle_file.config['spider_status'] == 3:
            print(time.strftime("wait for resume command %Y-%m-%d %H:%M:%S", time.localtime()))
            time.sleep(1)
